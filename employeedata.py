#NB: ALL files must be in the same directory
import openpyxl
def Excel_file():
    # Get file and sheet name from user input
    file = input("Enter filename with its extension: ")
    pb = openpyxl.load_workbook(file)
    sheet = input("Enter sheet name:  ")
    ws = pb[sheet]  
    # Get strings to be updated
    string_replace = input("Enter the string to be replace:  ")
    new_string = input("Enter new string to replace:  ")
    # Keep track of the number of cells updated
    i = 0 
# Iterate over columns and rows to find word to replace
    for r in range(1,ws.max_row+1):
        for c in range(1,ws.max_column+1):
        # Convert the value of the column and row to string before checking for word to replace
            s = str(ws.cell(r,c).value)
            # Check if word is present, and then replaces each word with the new word
            if string_replace in s: 
                ws.cell(r,c).value = s.replace(string_replace,new_string) 
            # Print the cells to be formatted while incrementing i till the last row or column
                print("row {} col {} : {}".format(r,c,s))
                i += 1
    # Print number of updated cells
    print("{} cells updated for excel files".format(i))
    # Create a new file and save changes to it
    save_as = input("Enter new filename to save changes to:  ") 
    pb.save(save_as)


def Csv_file():
    # Get file and sheet name from user input
    file = input("Enter filename with its extension: ")
    text = open(file, "r")
     
    # Get strings to be updated
    string_replace = input("Enter the string to be replace:  ")
    new_string = input("Enter new string to replace:  ")
    # create and return a new string, then replaces the old word with the new
    
    text = ''.join([i for i in text]) \
    .replace(string_replace, new_string)
    # Takes filename from user to save changes to
    save_as = input("Enter filename to save changes to: ")
    x = open(save_as,"w")                   # opens a new file to save changes to
    x.writelines(text)                      # write changes to the new file
    x.close()
    
class main:
# Checks type of input and call the correspinding function
    file_type = input("Enter file type - csv or excel:  ")
    if file_type=="excel":
        Excel_file()
    elif file_type == "csv":
        Csv_file()
# Run again after exit
main()